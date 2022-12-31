
#link
#https://www.youtube.com/watch?v=PqiwVp53ftU&list=FL6BJE5AMHOkFmZSdfKEyq0w&index=5&ab_channel=ThePyCoach
#pagina oficial pandas
#https://pandas.pydata.org/
#lista de funciones aggfunc de pandas
#https://datascientyst.com/list-aggregation-functions-aggfunc-groupby-pandas/

#---------------------------------------------------------------------------------------
# %%
#Generar un fichero excel que contiene una tabla pivote a partir de otro fichero excel dado mediante la libre pandas

#Tratar datos del excel, seleccionar y filtrar 
import pandas as pd

#lee archivo excel con pandas
archivo_excel=pd.read_excel('supermarket_sales.xlsx')

#lee archivo excel con pandas
#print(archivo_excel[['Gender','Product line','Total']])#muestra fichero

#tabla pivote agrupa los datos que queremos extraer(cuantos gastan mujeres y hombres, que tipos de productos y cuanto es la cantidad)
# index->indice, columns->columnas,values->valores de la suma aggfunc='sum'->funcion sumar,round(0)-> sin decimales
tabla_pivote=archivo_excel.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)
print(tabla_pivote)

#exportar tabla pivote en excel
#startrow->fila donde empieza a dibujar tabla, sheet_name->nombre de la hoja del excel
tabla_pivote.to_excel('sales_2021.xlsx',startrow=5, sheet_name='Report')

#--------------------------------------------------------------------------------
#-----------------------------------------
# %%
#Como Obtener la posicion de la tabla


#editar excel en python
from openpyxl import load_workbook

#lee archivo excel con openpyxl
wb=load_workbook('sales_2021.xlsx')
pestaña= wb['Report']

#obtener posicion de la tabla
min_col=wb.active.min_column#colum donde empieza la tabla
max_col=wb.active.max_column#colum donde finaliza la tabla
min_fila=wb.active.min_row#fila donde empieza la tabla
max_fila=wb.active.max_row#fila donde finaliza la tabla

print(min_col)#1
print(max_col)#7
print(min_fila)#6
print(max_fila)#8
#---------------------------------------------------------------------------------------

#-----------------------------------------
# %%

#Dibujar Grafico en un archivo Excel
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

#Hacer reportes en Excel con Opengpyxl
wb=load_workbook('sales_2021.xlsx')
pestaña= wb['Report']

#obtener posicion de la tabla
min_col=wb.active.min_column#1(A)
max_col=wb.active.max_column#7(G)
min_fila=wb.active.min_row #6
max_fila=wb.active.max_row#8

#crear un Grafico
barchart=BarChart()

#obtener region de la tabla(columnas B-G y filas 6-8) para grafico eje y
data = Reference(pestaña, min_col=min_col+1,max_col=max_col,min_row=min_fila,max_row=max_fila)
#obtener region de la tabla(columnas A-A, filas 7-8) para grafico eje x
categorias = Reference(pestaña, min_col=min_col,max_col=min_col,min_row=min_fila+1,max_row=max_fila)

#incluyendo los datos al grafico
barchart.add_data(data,titles_from_data=True)
#incluyendo las categorias al grafico
barchart.set_categories(categorias)

#agregar barchart a la pestaña, y la celda donde empieze a dibujarse el grafico
pestaña.add_chart(barchart,'B12')
#agregar el titulo al grafico
barchart.title='Ventas'
#agregar estilo al grafico
barchart.style=5
#guardar el archivo
wb.save('sales_2021.xlsx')
# %%
